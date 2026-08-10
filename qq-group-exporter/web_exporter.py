#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
QQ群成员导出工具 - Web版
连接到 NapCatQQ 的 HTTP API (OneBot v11)
"""
import os
import requests
from datetime import datetime, date
from pathlib import Path

from flask import Flask, jsonify, send_file, request
from flask_cors import CORS

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


app = Flask(__name__)
CORS(app)

EXPORT_DIR = Path("exports")
EXPORT_DIR.mkdir(exist_ok=True)

ONEBOT_URL = os.environ.get("ONEBOT_URL", "http://127.0.0.1:8081")


def export_to_excel(group_id, group_name, members):
    """导出群成员到Excel - 按入群时间正序，列：QQ号/昵称/群名片/角色/性别/入群时间/最后发言"""
    if not OPENPYXL_AVAILABLE:
        return None, "openpyxl 未安装，请运行：pip install openpyxl"

    members_sorted = sorted(members, key=lambda m: m.get("join_time", 0))

    wb = openpyxl.Workbook()
    ws = wb.active
    wb.remove(ws)
    ws = wb.create_sheet("成员名单")

    header_font = Font(name="微软雅黑", size=11, bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    headers = ["QQ号", "昵称", "群名片", "角色", "入群时间", "最后发言"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = thin_border

    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(horizontal="left", vertical="center")

    role_map = {"owner": "群主", "admin": "管理员", "member": "成员"}

    for row_num, member in enumerate(members_sorted, 2):
        user_id = member.get("user_id", "")
        nickname = member.get("nickname", "")
        card = member.get("card", "")
        display_name = card if card else nickname
        role = role_map.get(member.get("role", "member"), "成员")

        ws.cell(row=row_num, column=1, value=str(user_id))
        ws.cell(row=row_num, column=2, value=nickname)
        ws.cell(row=row_num, column=3, value=display_name)
        ws.cell(row=row_num, column=4, value=role)

        join_time = member.get("join_time")
        if join_time:
            ws.cell(row=row_num, column=5, value=datetime.fromtimestamp(join_time).strftime("%Y-%m-%d %H:%M:%S"))

        last_sent = member.get("last_sent_time")
        if last_sent:
            ws.cell(row=row_num, column=6, value=datetime.fromtimestamp(last_sent).strftime("%Y-%m-%d %H:%M:%S"))

        for col_num in range(1, 7):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    safe_name = "".join(c for c in group_name if c not in r'\/:*?"<>|')
    filename = EXPORT_DIR / f"{safe_name}-{group_id}-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"

    wb.save(filename)
    return filename, None


def call_onebot_api(action, params=None):
    """调用 OneBot API"""
    try:
        resp = requests.post(
            f"{ONEBOT_URL}/{action}",
            json=params or {},
            timeout=60
        )
        result = resp.json()
        if result.get("retcode") != 0:
            return None, result.get("msg", "API调用失败")
        return result.get("data"), None
    except requests.exceptions.ConnectionError:
        return None, "无法连接到 OneBot API，请确认 NapCatQQ 正在运行"
    except Exception as e:
        return None, str(e)


def check_login():
    data, error = call_onebot_api("get_login_info")
    return error is None


def get_groups():
    """获取群列表（按群主入群时间即群创建时间正序，带群主信息）"""
    data, error = call_onebot_api("get_group_list")
    if error:
        return [], error

    groups = []
    for group in data:
        gid = group.get("group_id")
        owner = None
        create_time = 0
        members_data, err = call_onebot_api("get_group_member_list", {"group_id": gid})
        if not err and isinstance(members_data, list):
            for m in members_data:
                if m.get("role") == "owner":
                    owner = m
                    create_time = m.get("join_time", 0)
                    break

        groups.append({
            "group_id": gid,
            "group_name": group.get("group_name"),
            "member_count": group.get("member_count"),
            "max_member_count": group.get("max_member_count"),
            "avatar_url": f"https://p.qlogo.cn/gh/{gid}/{gid}/100",
            "owner_id": owner.get("user_id") if owner else None,
            "owner_nickname": owner.get("nickname") if owner else "",
            "owner_avatar": f"https://q1.qlogo.cn/g?b=qq&nk={owner.get('user_id')}&s=100" if owner else "",
            "create_time": create_time,
        })

    groups.sort(key=lambda g: g["create_time"] or 0, reverse=True)
    return groups, None


def get_group_members(group_id):
    data, error = call_onebot_api(
        "get_group_member_list",
        {"group_id": group_id, "no_cache": True}
    )
    if error:
        return [], error
    return data, None


@app.route("/")
def index():
    return send_file("templates/index.html")


@app.route("/stats")
def stats_page():
    return send_file("templates/stats.html")


@app.route("/api/status")
def get_status():
    logged_in = check_login()
    return jsonify({
        "logged_in": logged_in,
        "login_status": "已连接" if logged_in else "未连接，请先启动 NapCatQQ",
    })


@app.route("/api/groups")
def get_groups_api():
    if not check_login():
        return jsonify({"error": "未连接到 OneBot API"}), 401

    groups, error = get_groups()
    if error:
        return jsonify({"error": error}), 500

    return jsonify({"groups": groups, "total": len(groups)})


@app.route("/api/export/<int:group_id>", methods=["POST"])
def export_group_api(group_id):
    if not check_login():
        return jsonify({"error": "未连接到 OneBot API"}), 401

    groups, error = get_groups()
    if error:
        return jsonify({"error": error}), 500

    group_name = "群"
    for g in groups:
        if g.get("group_id") == group_id:
            group_name = g.get("group_name", "群")
            break

    members, error = get_group_members(group_id)
    if error:
        return jsonify({"error": error}), 500

    filename, error = export_to_excel(group_id, group_name, members)
    if error:
        return jsonify({"error": error}), 500

    return send_file(filename, as_attachment=True)


def get_friends():
    """获取好友列表"""
    data, error = call_onebot_api("get_friend_list")
    if error:
        return [], error
    return data, None


def get_managed_groups():
    """获取当前登录用户为群主或管理员的群，按创建时间从旧到新排序"""
    groups, error = get_groups()
    if error:
        return [], error

    login_info, err = call_onebot_api("get_login_info")
    if err or not login_info:
        return [], err or "无法获取登录信息"

    self_id = login_info.get("user_id")

    managed = []
    for g in groups:
        gid = g["group_id"]
        members, err = get_group_members(gid)
        if err:
            continue
        for m in members:
            if m.get("user_id") == self_id and m.get("role") in ("owner", "admin"):
                managed.append(g)
                break

    managed.sort(key=lambda g: g.get("create_time") or 0)
    return managed, None


def export_stats_to_excel(main_group_id, main_group_name, scanner_users, water_group_ids,
                          black_group_ids, staff_user_ids, optional_group_ids,
                          start_date_str, end_date_str):
    """
    参团率统计导出Excel
    主群提供用户池（排除群主/管理员）
    每个扫者一列，统计主群成员在该扫者为群主的群中的出现次数
    时间范围：扫者群的创建时间在 [start_date, end_date] 内
    主群和自选群不受时间范围限制，强制参与统计
    排除水群
    排除黑群：黑群的所有成员从用户池中移除
    排除工作人员：工作人员从用户池中移除
    """
    if not OPENPYXL_AVAILABLE:
        return None, "openpyxl 未安装，请运行：pip install openpyxl"

    start_ts = int(datetime.strptime(start_date_str, "%Y-%m-%d").timestamp())
    end_ts = int(datetime.strptime(end_date_str + " 23:59:59", "%Y-%m-%d %H:%M:%S").timestamp())

    water_set = set(water_group_ids)
    black_set = set(black_group_ids or [])
    optional_set = set(optional_group_ids or [])

    # 1. 获取主群成员列表（排除群主和管理员）
    main_members, error = get_group_members(main_group_id)
    if error:
        return None, f"获取主群成员失败: {error}"

    user_pool = [m for m in main_members if m.get("role") == "member"]
    if not user_pool:
        return None, "主群中没有普通成员（已排除群主和管理员）"

    # 1.1 排除黑群中的所有成员
    if black_set:
        black_member_ids = set()
        for bgid in black_set:
            bg_members, err = get_group_members(bgid)
            if err:
                continue
            for m in bg_members:
                black_member_ids.add(m.get("user_id"))
        user_pool = [m for m in user_pool if m["user_id"] not in black_member_ids]
        if not user_pool:
            return None, "排除黑群后没有剩余成员"

    # 1.2 排除工作人员
    if staff_user_ids:
        staff_set = set(staff_user_ids)
        user_pool = [m for m in user_pool if m["user_id"] not in staff_set]
        if not user_pool:
            return None, "排除工作人员后没有剩余成员"

    user_map = {m["user_id"]: m for m in user_pool}

    # 2. 获取所有群列表，找出每位扫者为群主的群
    all_groups_list, error = get_groups()
    if error:
        return None, f"获取群列表失败: {error}"

    group_map = {g["group_id"]: g for g in all_groups_list}

    scanner_group_map = {}
    for scanner in scanner_users:
        scanner_id = scanner["user_id"]
        scanner_groups = []
        for g in all_groups_list:
            if g["group_id"] in water_set:
                continue
            ct = g.get("create_time") or 0
            if ct < start_ts or ct > end_ts:
                continue
            if g.get("owner_id") == scanner_id:
                scanner_groups.append(g)
        scanner_group_map[scanner_id] = scanner_groups

    # 2.1 主群和自选群强制加入对应扫者的群列表（不受时间范围限制，排除水群）
    forced_group_ids = {main_group_id} | optional_set
    for gid in forced_group_ids:
        if gid in water_set:
            continue
        g = group_map.get(gid)
        if not g:
            continue
        owner_id = g.get("owner_id")
        if not owner_id or owner_id not in scanner_group_map:
            continue
        if not any(sg["group_id"] == gid for sg in scanner_group_map[owner_id]):
            scanner_group_map[owner_id].append(g)

    # 3. 对每位扫者的每个群，拉成员并统计主群成员出现次数
    stats = {uid: {scanner["user_id"]: 0 for scanner in scanner_users} for uid in user_map}

    for scanner in scanner_users:
        scanner_id = scanner["user_id"]
        for g in scanner_group_map[scanner_id]:
            members, err = get_group_members(g["group_id"])
            if err:
                continue
            for m in members:
                uid = m.get("user_id")
                if uid in stats:
                    stats[uid][scanner_id] += 1

    # 4. 汇总排序
    rows = []
    for uid, scanner_counts in stats.items():
        total = sum(scanner_counts.values())
        rows.append({
            "user_id": uid,
            "nickname": user_map[uid].get("nickname", ""),
            "total": total,
            "scanner_counts": scanner_counts,
        })

    rows.sort(key=lambda r: r["total"], reverse=True)

    # 5. 收集团书群列表（按建群日期正序），用于右侧对照表
    scanner_groups_list = []
    for scanner in scanner_users:
        scanner_id = scanner["user_id"]
        scanner_name = scanner.get("nickname") or str(scanner_id)
        for g in scanner_group_map.get(scanner_id, []):
            scanner_groups_list.append({
                "scanner_name": scanner_name,
                "group_name": g.get("group_name", ""),
                "member_count": g.get("member_count", 0),
                "create_time": g.get("create_time") or 0,
            })
    scanner_groups_list.sort(key=lambda x: x["create_time"])

    # 6. 写入Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    wb.remove(ws)
    ws = wb.create_sheet("参团率统计")

    header_font = Font(name="微软雅黑", size=11, bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(horizontal="left", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")

    headers = ["QQ号", "昵称", "总参团次数"]
    for scanner in scanner_users:
        display = scanner.get("remark") or scanner.get("nickname") or str(scanner["user_id"])
        headers.append(display)

    scanner_col_start = len(headers) + 2  # 空1列
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # 右侧扫者群对照表表头
    right_headers = ["扫者ID", "团书群", "人数", "建群日期"]
    for i, h in enumerate(right_headers):
        cell = ws.cell(row=1, column=scanner_col_start + i, value=h)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = thin_border

    for row_num, row_data in enumerate(rows, 2):
        ws.cell(row=row_num, column=1, value=str(row_data["user_id"]))
        ws.cell(row=row_num, column=2, value=row_data["nickname"])
        ws.cell(row=row_num, column=3, value=row_data["total"])

        for col_idx, scanner in enumerate(scanner_users, 4):
            ws.cell(row=row_num, column=col_idx, value=row_data["scanner_counts"][scanner["user_id"]])

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            if col_num >= 3:
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment

    # 右侧扫者群对照表数据
    for i, sg in enumerate(scanner_groups_list):
        r = 2 + i
        ws.cell(row=r, column=scanner_col_start, value=sg["scanner_name"])
        ws.cell(row=r, column=scanner_col_start + 1, value=sg["group_name"])
        ws.cell(row=r, column=scanner_col_start + 2, value=sg["member_count"])
        if sg["create_time"]:
            ws.cell(row=r, column=scanner_col_start + 3,
                    value=datetime.fromtimestamp(sg["create_time"]).strftime("%Y-%m-%d"))
        for j in range(4):
            cell = ws.cell(row=r, column=scanner_col_start + j)
            cell.font = data_font
            cell.border = thin_border
            if j in (2, 3):
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    for i in range(len(scanner_users)):
        col_letter = openpyxl.utils.get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 14

    # 空列宽度 + 右侧四列宽度
    blank_col = openpyxl.utils.get_column_letter(scanner_col_start - 1)
    ws.column_dimensions[blank_col].width = 3
    ws.column_dimensions[openpyxl.utils.get_column_letter(scanner_col_start)].width = 14
    ws.column_dimensions[openpyxl.utils.get_column_letter(scanner_col_start + 1)].width = 20
    ws.column_dimensions[openpyxl.utils.get_column_letter(scanner_col_start + 2)].width = 8
    ws.column_dimensions[openpyxl.utils.get_column_letter(scanner_col_start + 3)].width = 14

    ws.freeze_panes = "A2"

    def _fmt_date(s):
        d = datetime.strptime(s, "%Y-%m-%d")
        return f"{d.year}.{d.month}.{d.day}"
    start_fmt = _fmt_date(start_date_str)
    end_fmt = _fmt_date(end_date_str)
    safe_name = "".join(c for c in main_group_name if c not in r'\/:*?"<>|')
    filename = EXPORT_DIR / f"参团率统计（{start_fmt}-{end_fmt}）-{safe_name}.xlsx"

    wb.save(filename)
    return filename, None


@app.route("/api/stats/managed-groups")
def stats_managed_groups():
    if not check_login():
        return jsonify({"error": "未连接到 OneBot API"}), 401
    groups, error = get_managed_groups()
    if error:
        return jsonify({"error": error}), 500
    return jsonify({"managed_groups": groups, "total": len(groups)})


@app.route("/api/stats/friends")
def stats_friends():
    if not check_login():
        return jsonify({"error": "未连接到 OneBot API"}), 401
    friends, error = get_friends()
    if error:
        return jsonify({"error": error}), 500
    return jsonify({"friends": friends, "total": len(friends)})


@app.route("/api/stats/export", methods=["POST"])
def stats_export():
    if not check_login():
        return jsonify({"error": "未连接到 OneBot API"}), 401

    data = request.get_json(force=True)
    main_group_id = data.get("main_group_id")
    scanner_users_raw = data.get("scanner_users", [])
    water_group_ids = data.get("water_group_ids", [])
    black_group_ids = data.get("black_group_ids", [])
    staff_user_ids = data.get("staff_user_ids", [])
    optional_group_ids = data.get("optional_group_ids", [])
    start_date = data.get("start_date")
    end_date = data.get("end_date")

    if not main_group_id:
        return jsonify({"error": "请选择主群"}), 400
    if not scanner_users_raw:
        return jsonify({"error": "请至少选择一位扫者"}), 400
    if not start_date or not end_date:
        return jsonify({"error": "请选择时间范围"}), 400

    # 按 user_id 去重
    seen = set()
    scanner_users = []
    for u in scanner_users_raw:
        uid = u.get("user_id")
        if uid and uid not in seen:
            seen.add(uid)
            scanner_users.append(u)

    # 获取主群信息
    groups, error = get_groups()
    if error:
        return jsonify({"error": error}), 500

    main_group_name = "主群"
    for g in groups:
        if g.get("group_id") == main_group_id:
            main_group_name = g.get("group_name", "主群")
            break

    filename, error = export_stats_to_excel(
        main_group_id, main_group_name, scanner_users, water_group_ids,
        black_group_ids, staff_user_ids, optional_group_ids, start_date, end_date
    )
    if error:
        return jsonify({"error": error}), 500

    return send_file(filename, as_attachment=True, download_name=filename.name)


def main():
    print("=" * 66)
    print("                 QQ群成员导出工具")
    print("=" * 66)
    print()
    print("使用步骤：")
    print("1. 启动 NapCatQQ 并确保 QQ 已登录")
    print(f"2. 确保 HTTP API 服务运行在 {ONEBOT_URL}")
    print("3. 打开浏览器访问：http://127.0.0.1:8080")
    print()
    print("=" * 66)

    app.run(host="0.0.0.0", port=8080, debug=True)


if __name__ == "__main__":
    main()
